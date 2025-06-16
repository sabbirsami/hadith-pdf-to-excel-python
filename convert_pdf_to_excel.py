import os
import cv2
import numpy as np
from PIL import Image
import pytesseract
import fitz  # PyMuPDF
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

class FixedPDFToExcelConverter:
    def __init__(self, pdf_path, output_dir="output"):
        self.pdf_path = pdf_path
        self.output_dir = output_dir
        self.images_dir = os.path.join(output_dir, "images")
        self.processed_images_dir = os.path.join(output_dir, "processed_images")
        self.ocr_text_dir = os.path.join(output_dir, "ocr_text")

        # Create directories
        for dir_path in [self.output_dir, self.images_dir, self.processed_images_dir, self.ocr_text_dir]:
            os.makedirs(dir_path, exist_ok=True)

    def find_first_chapter_page(self):
        """Find the page number where first অধ্যায়: appears"""
        print("Finding first chapter page...")

        doc = fitz.open(self.pdf_path)

        for page_num in range(len(doc)):
            page = doc.load_page(page_num)
            text = page.get_text()

            if 'অধ্যায়:' in text or 'অধ্যায়ঃ' in text:
                print(f"Found first অধ্যায়: on page {page_num + 1}")
                doc.close()
                return page_num

        doc.close()
        print("No অধ্যায়: found, starting from page 1")
        return 0

    def step1_pdf_to_images(self, start_page, max_pages=20):
        """Convert PDF pages to images starting from first chapter page"""
        print(f"Step 1: Converting PDF to images starting from page {start_page + 1}...")

        doc = fitz.open(self.pdf_path)
        end_page = min(start_page + max_pages, len(doc))

        for page_num in range(start_page, end_page):
            page = doc.load_page(page_num)
            mat = fitz.Matrix(2.0, 2.0)  # High resolution
            pix = page.get_pixmap(matrix=mat)
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)

            # Save image with sequential numbering
            img_path = os.path.join(self.images_dir, f"page_{page_num - start_page + 1:03d}.png")
            img.save(img_path, "PNG")
            print(f"Saved: page_{page_num - start_page + 1:03d}.png (original page {page_num + 1})")

        doc.close()
        print(f"Converted {end_page - start_page} pages to images")

    def step3_remove_arabic_and_extras(self):
        """Remove Arabic text, headers, and page numbers"""
        print("Step 3: Removing Arabic text and extra information...")

        for img_file in sorted(os.listdir(self.images_dir)):
            if img_file.endswith('.png'):
                img_path = os.path.join(self.images_dir, img_file)
                image = Image.open(img_path)
                img_array = np.array(image)

                height, width = img_array.shape[:2]

                # Create mask to remove unwanted areas
                mask = np.ones((height, width), dtype=np.uint8) * 255

                # Remove top 12% (header/title area)
                mask[0:int(height*0.12), :] = 0

                # Remove bottom 10% (page numbers and references)
                mask[int(height*0.90):, :] = 0

                # Apply mask - set masked areas to white
                processed_img = img_array.copy()
                processed_img[mask == 0] = [255, 255, 255]

                # Additional Arabic text removal using contour detection
                gray = cv2.cvtColor(processed_img, cv2.COLOR_RGB2GRAY)
                _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

                contours, _ = cv2.findContours(binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

                # Remove Arabic text blocks (typically wider horizontal blocks)
                for contour in contours:
                    x, y, w, h = cv2.boundingRect(contour)
                    # Arabic text characteristics: wide horizontal blocks
                    if w > 300 and h > 20 and w/h > 8:
                        processed_img[y:y+h, x:x+w] = [255, 255, 255]

                # Save processed image
                processed_image = Image.fromarray(processed_img)
                output_path = os.path.join(self.processed_images_dir, img_file)
                processed_image.save(output_path)
                print(f"Processed: {img_file}")

        print("Arabic text and extras removed from all images")

    def step4_ocr_images(self):
        """Perform OCR on processed images"""
        print("Step 4: Performing OCR on images...")

        all_text = []

        for img_file in sorted(os.listdir(self.processed_images_dir)):
            if img_file.endswith('.png'):
                img_path = os.path.join(self.processed_images_dir, img_file)

                # Configure OCR for Bengali
                custom_config = r'--oem 3 --psm 6 -l ben+eng'

                try:
                    text = pytesseract.image_to_string(Image.open(img_path), config=custom_config)

                    # Save individual OCR result
                    text_file = os.path.join(self.ocr_text_dir, f"{img_file.replace('.png', '.txt')}")
                    with open(text_file, 'w', encoding='utf-8') as f:
                        f.write(text)

                    all_text.append(f"=== PAGE {img_file} ===\n{text}\n")
                    print(f"OCR completed: {img_file}")

                except Exception as e:
                    print(f"OCR failed for {img_file}: {str(e)}")
                    all_text.append(f"=== PAGE {img_file} ===\n[OCR FAILED]\n")

        # Save combined OCR text
        combined_text_path = os.path.join(self.output_dir, "combined_ocr_text.txt")
        with open(combined_text_path, 'w', encoding='utf-8') as f:
            f.writelines(all_text)

        print("OCR completed for all images")
        return combined_text_path

    def is_bengali_text(self, text):
        """Check if text contains Bengali characters"""
        if not text:
            return False
        bengali_pattern = r'[\u0980-\u09FF]'
        return bool(re.search(bengali_pattern, text))

    def clean_hadith_text(self, text):
        """Clean hadith text by removing references and extra content"""
        if not text:
            return ""

        # Remove common reference patterns
        reference_patterns = [
            r'সহিহুল বুখারি[^।]*।?',
            r'সহিহ মুসলিম[^।]*।?',
            r'সুনানু তিরমিযি[^।]*।?',
            r'সুনানু আবু দাউদ[^।]*।?',
            r'সুনানু নাসাই[^।]*।?',
            r'সুনানু ইবনে মাজাহ[^।]*।?',
            r'মুসনাদে আহমাদ[^।]*।?',
            r'[০-৯]+\s*[।.\-\s]*\s*[০-৯]*',  # Remove numbers and punctuation
            r'[a-zA-Z0-9\s]{10,}',  # Remove long English/number sequences
            r'[^\u0980-\u09FF\s।,;:\-\[\]]+',  # Remove non-Bengali except basic punctuation
        ]

        cleaned_text = text
        for pattern in reference_patterns:
            cleaned_text = re.sub(pattern, '', cleaned_text)

        # Clean up extra spaces and punctuation
        cleaned_text = re.sub(r'\s+', ' ', cleaned_text)
        cleaned_text = re.sub(r'[।]{2,}', '।', cleaned_text)
        cleaned_text = cleaned_text.strip()

        return cleaned_text

    def extract_data_from_text(self, text_file_path):
        """Extract chapters, sections, and hadiths from OCR text"""
        print("Extracting data from OCR text...")

        with open(text_file_path, 'r', encoding='utf-8') as f:
            content = f.read()

        chapters = []
        sections = []
        hadiths = []

        chapter_id = 1
        section_id = 1
        current_chapter_context = False

        # Split content into lines
        lines = content.split('\n')

        i = 0
        while i < len(lines):
            line = lines[i].strip()

            if not line:
                i += 1
                continue

            # Skip page markers
            if 'PAGE' in line or '===' in line:
                i += 1
                continue

            # Extract chapters - look for অধ্যায়: or অধ্যায়ঃ
            chapter_patterns = [r'অধ্যায়:\s*(.+)', r'অধ্যায়ঃ\s*(.+)']
            chapter_found = False

            for pattern in chapter_patterns:
                match = re.search(pattern, line)
                if match:
                    chapter_name = match.group(1).strip()
                    # Clean chapter name
                    chapter_name = re.sub(r'^[\d\s\-।\.]+', '', chapter_name)
                    chapter_name = re.sub(r'\s+', ' ', chapter_name).strip()

                    if chapter_name and self.is_bengali_text(chapter_name):
                        chapters.append({
                            'id': chapter_id,
                            'name': chapter_name
                        })
                        print(f"Found chapter {chapter_id}: {chapter_name}")
                        chapter_id += 1
                        current_chapter_context = True
                        chapter_found = True
                        break

            if chapter_found:
                i += 1
                continue

            # Extract sections - Bengali text that appears after chapters
            if (current_chapter_context and
                self.is_bengali_text(line) and
                not re.match(r'^\[\d+\]', line) and  # Not a hadith
                len(line.split()) >= 3 and len(line.split()) <= 15 and  # Reasonable length
                not line.endswith('।') and  # Not sentence-like
                'সহিহ' not in line and 'মুসলিম' not in line):  # Not references

                sections.append({
                    'id': section_id,
                    'name': line
                })
                print(f"Found section {section_id}: {line}")
                section_id += 1
                i += 1
                continue

            # Extract hadiths - text starting with [number]
            hadith_match = re.match(r'^\[(\d+)\]\s*(.*)', line)
            if hadith_match:
                hadith_id = hadith_match.group(1)
                hadith_text = hadith_match.group(2).strip()

                # Collect continuation lines
                j = i + 1
                while j < len(lines):
                    next_line = lines[j].strip()

                    if not next_line:
                        j += 1
                        continue

                    # Stop conditions
                    if (re.match(r'^\[\d+\]', next_line) or  # Next hadith
                        'অধ্যায়:' in next_line or 'অধ্যায়ঃ' in next_line or  # Next chapter
                        'PAGE' in next_line or '===' in next_line or  # Page marker
                        'সহিহ' in next_line):  # References start
                        break

                    # Add Bengali text
                    if self.is_bengali_text(next_line):
                        hadith_text += " " + next_line

                    j += 1

                # Clean and save hadith
                hadith_text = self.clean_hadith_text(hadith_text)

                if hadith_text and len(hadith_text) > 20:
                    hadiths.append({
                        'id': hadith_id,
                        'hadith': hadith_text
                    })
                    print(f"Found hadith [{hadith_id}]: {hadith_text[:50]}...")

                i = j
                continue

            i += 1

        print(f"Extracted: {len(chapters)} chapters, {len(sections)} sections, {len(hadiths)} hadiths")
        return chapters, sections, hadiths

    def step6_create_excel(self, chapters, sections, hadiths):
        """Create Excel file with three sheets"""
        print("Step 6: Creating Excel file...")

        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)

        # Create Chapter sheet
        chapter_sheet = wb.create_sheet(title="chapter")
        chapter_sheet['A1'] = 'id'
        chapter_sheet['B1'] = 'name'

        # Format headers
        for cell in ['A1', 'B1']:
            chapter_sheet[cell].font = Font(bold=True)
            chapter_sheet[cell].alignment = Alignment(horizontal='center')

        # Add chapter data
        for i, chapter in enumerate(chapters, 2):
            chapter_sheet[f'A{i}'] = chapter['id']
            chapter_sheet[f'B{i}'] = chapter['name']
            chapter_sheet[f'A{i}'].alignment = Alignment(horizontal='center')
            chapter_sheet[f'B{i}'].alignment = Alignment(horizontal='left')

        # Create Section sheet
        section_sheet = wb.create_sheet(title="section")
        section_sheet['A1'] = 'id'
        section_sheet['B1'] = 'name'

        # Format headers
        for cell in ['A1', 'B1']:
            section_sheet[cell].font = Font(bold=True)
            section_sheet[cell].alignment = Alignment(horizontal='center')

        # Add section data
        for i, section in enumerate(sections, 2):
            section_sheet[f'A{i}'] = section['id']
            section_sheet[f'B{i}'] = section['name']
            section_sheet[f'A{i}'].alignment = Alignment(horizontal='center')
            section_sheet[f'B{i}'].alignment = Alignment(horizontal='left')

        # Create Hadith sheet
        hadith_sheet = wb.create_sheet(title="hadith")
        hadith_sheet['A1'] = 'id'
        hadith_sheet['B1'] = 'hadith'

        # Format headers
        for cell in ['A1', 'B1']:
            hadith_sheet[cell].font = Font(bold=True)
            hadith_sheet[cell].alignment = Alignment(horizontal='center')

        # Add hadith data
        for i, hadith in enumerate(hadiths, 2):
            hadith_sheet[f'A{i}'] = hadith['id']
            hadith_sheet[f'B{i}'] = hadith['hadith']
            hadith_sheet[f'A{i}'].alignment = Alignment(horizontal='center')
            hadith_sheet[f'B{i}'].alignment = Alignment(horizontal='left')

        # Adjust column widths
        for sheet in [chapter_sheet, section_sheet, hadith_sheet]:
            sheet.column_dimensions['A'].width = 10
            sheet.column_dimensions['B'].width = 80

        # Save Excel file
        excel_path = os.path.join(self.output_dir, "hadith_data.xlsx")
        wb.save(excel_path)

        print(f"Excel file created: {excel_path}")
        print(f"Final count - Chapters: {len(chapters)}, Sections: {len(sections)}, Hadiths: {len(hadiths)}")
        return excel_path

    def run_conversion(self):
        """Run the complete conversion process"""
        print("Starting PDF to Excel conversion...")
        print("="*50)

        try:
            # Find first chapter page
            start_page = self.find_first_chapter_page()

            # Step 1: PDF to Images (20 pages from first chapter)
            self.step1_pdf_to_images(start_page, max_pages=20)

            # Step 3: Remove Arabic text and extras
            self.step3_remove_arabic_and_extras()

            # Step 4: OCR
            ocr_text_path = self.step4_ocr_images()

            # Extract data and create Excel
            chapters, sections, hadiths = self.extract_data_from_text(ocr_text_path)
            excel_path = self.step6_create_excel(chapters, sections, hadiths)

            print("="*50)
            print("Conversion completed successfully!")
            print(f"Started from page: {start_page + 1}")
            print(f"Processed 20 pages from first chapter")
            print(f"Output directory: {self.output_dir}")
            print(f"Excel file: {excel_path}")

            return excel_path

        except Exception as e:
            print(f"Error during conversion: {str(e)}")
            import traceback
            traceback.print_exc()
            raise

# Usage
if __name__ == "__main__":
    pdf_path = "bangla_hadith.pdf"  # Replace with your PDF path
    converter = FixedPDFToExcelConverter(pdf_path)

    excel_file = converter.run_conversion()
    print(f"\nConversion completed! Excel file: {excel_file}")
